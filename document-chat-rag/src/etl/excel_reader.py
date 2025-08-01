"""
Excel Reader Module

This module reads Excel files containing PDF download links.
According to the project requirements, column G holds the links to each PDF file.
"""

import os
from typing import List, Dict, Any
from openpyxl import load_workbook
from src.utils.logging import get_logger

logger = get_logger(__name__)

def read_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    Reads an Excel file and extracts PDF links from column G.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        List[Dict[str, Any]]: List of dictionaries containing PDF information
            Each dict contains:
            - url: The PDF download URL
            - filename: Extracted filename from URL or row index if not available
            - metadata: Other columns data as metadata
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    logger.info(f"Reading Excel file: {file_path}")
    
    try:
        # Load the workbook - IMPORTANT: read_only must be False to access hyperlinks
        workbook = load_workbook(file_path, read_only=False)
        worksheet = workbook.active
        
        # Get the maximum row and column count
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Check if there are at least 7 columns (A-G)
        if max_col < 7:
            raise ValueError("Excel file does not have enough columns. Expected at least 7 columns (A-G).")
        
        logger.info(f"Excel file loaded with {max_row} rows and {max_col} columns")
        
        # Extract PDF links from column G (7th column)
        pdf_data = []
        
        # Start from row 1, assuming first row might contain headers
        for row_idx in range(1, max_row + 1):
            # Get the cell in column G (7th column)
            cell = worksheet.cell(row=row_idx, column=7)
            
            # Try to get hyperlink target first
            pdf_url = None
            try:
                if cell.hyperlink and cell.hyperlink.target:
                    pdf_url = cell.hyperlink.target
                    logger.debug(f"Found hyperlink in row {row_idx}: {pdf_url}")
            except AttributeError:
                pass
            
            # If no hyperlink, try cell value
            if pdf_url is None:
                cell_value = cell.value
                # Skip empty cells
                if cell_value is None or str(cell_value).strip() == "":
                    logger.warning(f"Empty PDF URL in row {row_idx}, skipping")
                    continue
                
                # Convert to string in case it's not already
                pdf_url = str(cell_value).strip()
            
            # Skip if not a valid URL
            if not pdf_url.startswith(('http://', 'https://')):
                logger.warning(f"Invalid URL in row {row_idx}: {pdf_url}, skipping")
                continue
            
            # Extract filename from URL or use row index
            try:
                filename = pdf_url.split('/')[-1]
                # Ensure filename has .pdf extension
                if not filename.lower().endswith('.pdf'):
                    filename += '.pdf'
            except:
                filename = f"document_{row_idx}.pdf"
            
            # Collect other metadata from the row
            metadata = {}
            for col_idx in range(1, max_col + 1):
                if col_idx != 7:  # Skip column G which contains the URL
                    cell_val = worksheet.cell(row=row_idx, column=col_idx).value
                    # Use column letter as key (A, B, C, etc.)
                    col_letter = chr(ord('A') + col_idx - 1)
                    metadata[col_letter] = cell_val
            
            pdf_data.append({
                'url': pdf_url,
                'filename': filename,
                'metadata': metadata,
                'row_index': row_idx
            })
            
            logger.debug(f"Found PDF link: {pdf_url}")
        
        logger.info(f"Extracted {len(pdf_data)} valid PDF links from Excel file")
        return pdf_data
        
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        raise

def validate_excel_format(file_path: str) -> bool:
    """
    Validates that the Excel file has the expected format.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        bool: True if format is valid, False otherwise
    """
    try:
        workbook = load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        # Check if there are at least 7 columns (A-G)
        return worksheet.max_column >= 7
    except Exception as e:
        logger.error(f"Error validating Excel format: {str(e)}")
        return False

if __name__ == "__main__":
    # Example usage
    # This section can be used for testing
    print("Excel Reader Module")
    print("This module is designed to read Excel files containing PDF links in column G")